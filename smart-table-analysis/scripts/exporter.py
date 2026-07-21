#!/usr/bin/env python3
"""
Exporter - Smart Table Analysis Skill
Export analysis results to Excel, CSV, or formatted HTML report.
"""

import sys
import json
import argparse
from pathlib import Path
from datetime import datetime
import pandas as pd


def export_to_excel(df, output_path, sheet_name='原始数据', charts=None,
                     analysis_results=None, insights=None, summary_sheet='统计分析'):
    """Export DataFrame to Excel with multiple sheets:
    - Sheet1: Original data (auto-sized columns)
    - Sheet2: Statistical summary / analysis results (if provided)
    - Charts embedded in both sheets where applicable.
    """
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.utils import get_column_letter
    
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with pd.ExcelWriter(str(output_path), engine='openpyxl') as writer:
        # === Sheet 1: Original Data ===
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws_data = writer.sheets[sheet_name]
        
        # Auto-adjust column widths
        for i, col in enumerate(df.columns):
            max_len = max(
                df[col].astype(str).map(len).max(),
                len(str(col))
            )
            col_letter = get_column_letter(i + 1)
            ws_data.column_dimensions[col_letter].width = min(max_len + 2, 30)
        
        # Insert charts after data (if provided)
        if charts:
            start_row = len(df) + 3
            for i, chart in enumerate(charts):
                chart_path = chart.get('path', '')
                if not chart_path:
                    continue
                try:
                    img = XlImage(str(chart_path))
                    img.width = 600
                    img.height = 350
                    cell_ref = f'A{start_row + i * 20}'
                    ws_data.add_image(img, cell_ref)
                except Exception as e:
                    print(f'Warning: Could not insert chart {chart_path}: {e}', file=sys.stderr)
        
        # === Sheet 2: Statistical Summary ===
        if analysis_results or insights:
            # Build summary DataFrame
            summary_rows = []
            
            # Add insights as rows
            if insights:
                summary_rows.append(['=== 数据洞察 ===', '', ''])
                for insight in insights:
                    summary_rows.append([
                        insight.get('title', ''),
                        insight.get('message', ''),
                        insight.get('severity', '')
                    ])
                summary_rows.append(['', '', ''])
            
            # Add analysis results
            if analysis_results:
                summary_rows.append(['=== 统计分析 ===', '', ''])
                
                # Handle different result types
                if 'statistics' in analysis_results:
                    for col, stats in analysis_results['statistics'].items():
                        for stat_name, stat_val in stats.items():
                            summary_rows.append([f'{col}.{stat_name}', str(stat_val), ''])
                
                if 'group_column' in analysis_results:
                    summary_rows.append(['', '', ''])
                    summary_rows.append([
                        f'按 "{analysis_results["group_column"]}" 分组统计', '', ''
                    ])
                    if 'result' in analysis_results:
                        for group_name, group_data in analysis_results['result'].items():
                            if isinstance(group_data, dict):
                                for k, v in group_data.items():
                                    summary_rows.append([str(group_name), str(k), str(v)])
                            else:
                                summary_rows.append([str(group_name), str(group_data), ''])
                
                if 'correlation_matrix' in analysis_results:
                    summary_rows.append(['', '', ''])
                    summary_rows.append(['相关系数矩阵', '', ''])
                    corr = analysis_results['correlation_matrix']
                    for r_col, r_vals in corr.items():
                        for c_col, val in r_vals.items():
                            summary_rows.append([f'{r_col} vs {c_col}', f'{val:.4f}', ''])
                
                if 'strong_correlations' in analysis_results:
                    summary_rows.append(['', '', ''])
                    summary_rows.append(['强相关变量对', '', ''])
                    for item in analysis_results['strong_correlations']:
                        summary_rows.append([
                            f'{item["col_a"]} vs {item["col_b"]}',
                            f'r = {item["correlation"]}',
                            item['strength']
                        ])
                
                if 'outliers' in analysis_results:
                    summary_rows.append(['', '', ''])
                    summary_rows.append(['异常值检测', '', ''])
                    for col, info in analysis_results['outliers'].items():
                        summary_rows.append([
                            col,
                            f'异常值: {info.get("outlier_count", 0)}个 ({info.get("outlier_pct", 0)}%)',
                            f'范围: [{info.get("bounds", {}).get("lower", "")}, {info.get("bounds", {}).get("upper", "")}]'
                        ])
                
                if 'difference_insights' in analysis_results:
                    summary_rows.append(['', '', ''])
                    summary_rows.append(['组间差异分析', '', ''])
                    for diff in analysis_results['difference_insights']:
                        summary_rows.append([
                            diff.get('metric', ''),
                            f'最高: {diff.get("highest", "")} ({diff.get("highest_value", "")})',
                            f'最低: {diff.get("lowest", "")} ({diff.get("lowest_value", "")}) 差异: {diff.get("pct_difference", "")}%'
                        ])
            
            # Write summary sheet
            if summary_rows:
                summary_df = pd.DataFrame(summary_rows, columns=['项目', '值/指标', '备注'])
                summary_df.to_excel(writer, sheet_name=summary_sheet, index=False)
                ws_summary = writer.sheets[summary_sheet]
                # Auto-width
                for i, col in enumerate(summary_df.columns):
                    max_len = max(
                        summary_df[col].astype(str).map(len).max(),
                        len(str(col))
                    )
                    col_letter = get_column_letter(i + 1)
                    ws_summary.column_dimensions[col_letter].width = min(max_len + 2, 60)
    
    return {
        'export_path': str(output_path),
        'format': 'excel',
        'rows': len(df),
        'columns': len(df.columns),
        'sheets': [sheet_name, summary_sheet] if (analysis_results or insights) else [sheet_name],
        'charts_inserted': len(charts) if charts else 0
    }


def export_to_csv(df, output_path, encoding='utf-8-sig'):
    """Export DataFrame to CSV (with BOM for Excel compatibility)."""
    
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    df.to_csv(str(output_path), index=False, encoding=encoding)
    
    return {
        'export_path': str(output_path),
        'format': 'csv',
        'rows': len(df),
        'columns': len(df.columns)
    }


def generate_html_report(df, analysis_results, insights, charts, title='数据分析报告',
                         output_path=None):
    """Generate a beautifully formatted HTML analysis report with embedded chart images."""
    import base64
    
    timestamp = datetime.now().strftime('%Y年%m月%d日 %H:%M')
    
    # Build insights HTML
    insights_html = ''
    for insight in insights:
        severity_color = '#E74C3C' if insight.get('severity') == 'high' else '#F39C12' if insight.get('severity') == 'medium' else '#3498DB'
        insights_html += f'''
        <div class="insight-card" style="border-left: 4px solid {severity_color}">
            <div class="insight-title">{insight.get('title', '洞察')}</div>
            <div class="insight-body">{insight.get('message', '')}</div>
        </div>'''
    
    # Build stats HTML
    stats_html = ''
    if 'statistics' in analysis_results:
        for col, stats in analysis_results['statistics'].items():
            stats_html += f'''
            <div class="stat-card">
                <div class="stat-title">{col}</div>
                <div class="stat-items">
                    <div class="stat-item"><span class="stat-label">均值</span><span class="stat-value">{stats.get('mean', 'N/A')}</span></div>
                    <div class="stat-item"><span class="stat-label">中位数</span><span class="stat-value">{stats.get('50%', stats.get('median', 'N/A'))}</span></div>
                    <div class="stat-item"><span class="stat-label">标准差</span><span class="stat-value">{stats.get('std', 'N/A')}</span></div>
                    <div class="stat-item"><span class="stat-label">最小值</span><span class="stat-value">{stats.get('min', 'N/A')}</span></div>
                    <div class="stat-item"><span class="stat-label">最大值</span><span class="stat-value">{stats.get('max', 'N/A')}</span></div>
                    <div class="stat-item"><span class="stat-label">缺失率</span><span class="stat-value">{stats.get('missing_pct', 'N/A')}%</span></div>
                </div>
            </div>'''
    
    # Build charts HTML with embedded base64 images
    charts_html = ''
    for chart in charts:
        chart_path = chart.get('path', '')
        chart_title = chart.get('title', '')
        chart_desc = chart.get('description', '')
        if chart_path and Path(chart_path).exists():
            # Convert image to base64 for embedding
            with open(str(chart_path), 'rb') as f:
                img_data = base64.b64encode(f.read()).decode('utf-8')
            charts_html += f'''
            <div class="chart-card">
                <div class="chart-title">{chart_title}</div>
                <img src="data:image/png;base64,{img_data}" alt="{chart_title}" class="chart-image">
                <div class="chart-desc">{chart_desc}</div>
            </div>'''
        elif chart_path:
            charts_html += f'''
            <div class="chart-card">
                <div class="chart-title">{chart_title}</div>
                <div class="chart-desc">图表文件未找到: {chart_path}</div>
            </div>'''
    
    # Build data preview table
    preview_html = ''
    preview_df = df.head(20)
    headers = ''.join(f'<th>{col}</th>' for col in preview_df.columns)
    rows = ''
    for _, row in preview_df.iterrows():
        cells = ''.join(f'<td>{str(val)[:50]}</td>' for val in row)
        rows += f'<tr>{cells}</tr>'
    
    preview_html = f'''
    <table class="data-table">
        <thead><tr>{headers}</tr></thead>
        <tbody>{rows}</tbody>
    </table>'''
    
    html = f'''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Microsoft YaHei', 'PingFang SC', sans-serif; background: #f5f7fa; color: #2c3e50; line-height: 1.6; }
        .container { max-width: 960px; margin: 0 auto; padding: 20px; }
        .report-header { background: linear-gradient(135deg, #3498DB, #2ECC71); color: white; padding: 30px; border-radius: 12px; margin-bottom: 24px; }
        .report-header h1 { font-size: 24px; margin-bottom: 8px; }
        .report-header .meta { font-size: 14px; opacity: 0.8; }
        .section { background: white; border-radius: 12px; padding: 24px; margin-bottom: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }
        .section-title { font-size: 18px; font-weight: bold; margin-bottom: 16px; color: #2c3e50; border-bottom: 2px solid #3498DB; padding-bottom: 8px; }
        .insight-card { background: #f8f9fa; padding: 16px; margin-bottom: 12px; border-radius: 8px; }
        .insight-title { font-weight: bold; font-size: 14px; margin-bottom: 4px; }
        .insight-body { font-size: 13px; color: #555; }
        .stat-card { background: #f0f4f8; padding: 16px; border-radius: 8px; margin-bottom: 12px; }
        .stat-title { font-weight: bold; font-size: 15px; margin-bottom: 12px; color: #3498DB; }
        .stat-items { display: grid; grid-template-columns: repeat(3, 1fr); gap: 8px; }
        .stat-item { display: flex; justify-content: space-between; padding: 4px 8px; background: white; border-radius: 4px; font-size: 13px; }
        .stat-label { color: #7f8c8d; }
        .stat-value { font-weight: bold; color: #2c3e50; }
        .chart-card { background: white; padding: 16px; border-radius: 8px; margin-bottom: 16px; text-align: center; }
        .chart-title { font-weight: bold; margin-bottom: 12px; }
        .chart-image { max-width: 100%; border-radius: 8px; }
        .chart-desc { font-size: 13px; color: #7f8c8d; margin-top: 8px; }
        .data-table { width: 100%; border-collapse: collapse; font-size: 12px; }
        .data-table th { background: #3498DB; color: white; padding: 8px 12px; text-align: left; }
        .data-table td { padding: 6px 12px; border-bottom: 1px solid #ecf0f1; }
        .data-table tr:hover { background: #f8f9fa; }
        .quality-badge { display: inline-block; padding: 4px 12px; border-radius: 20px; font-size: 13px; font-weight: bold; }
        .quality-high { background: #27AE60; color: white; }
        .quality-medium { background: #F39C12; color: white; }
        .quality-low { background: #E74C3C; color: white; }
        .summary-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin-bottom: 20px; }
        .summary-item { background: white; padding: 16px; border-radius: 8px; text-align: center; box-shadow: 0 1px 4px rgba(0,0,0,0.1); }
        .summary-item .number { font-size: 24px; font-weight: bold; color: #3498DB; }
        .summary-item .label { font-size: 13px; color: #7f8c8d; margin-top: 4px; }
        @media (max-width: 768px) {
            .stat-items { grid-template-columns: repeat(2, 1fr); }
            .summary-grid { grid-template-columns: repeat(2, 1fr); }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="report-header">
            <h1>{title}</h1>
            <div class="meta">生成时间：{timestamp} | 数据行数：{len(df)} | 数据列数：{len(df.columns)}</div>
        </div>
        
        <div class="summary-grid">
            <div class="summary-item">
                <div class="number">{len(df)}</div>
                <div class="label">总行数</div>
            </div>
            <div class="summary-item">
                <div class="number">{len(df.columns)}</div>
                <div class="label">总列数</div>
            </div>
            <div class="summary-item">
                <div class="number">{df.select_dtypes(include=['number']).shape[1]}</div>
                <div class="label">数值列</div>
            </div>
            <div class="summary-item">
                <div class="number">{df.select_dtypes(exclude=['number']).shape[1]}</div>
                <div class="label">分类列</div>
            </div>
        </div>
        
        {f'<div class="section"><div class="section-title">数据洞察</div>{insights_html}</div>' if insights_html else ''}
        
        {f'<div class="section"><div class="section-title">统计摘要</div>{stats_html}</div>' if stats_html else ''}
        
        {f'<div class="section"><div class="section-title">可视化图表</div>{charts_html}</div>' if charts_html else ''}
        
        <div class="section">
            <div class="section-title">数据预览（前20行）</div>
            {preview_html}
        </div>
    </div>
</body>
</html>'''
    
    if output_path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(str(output_path), 'w', encoding='utf-8') as f:
            f.write(html)
        return {
            'export_path': str(output_path),
            'format': 'html',
            'rows': len(df),
            'columns': len(df.columns)
        }
    
    return {'html_content': html, 'format': 'html'}


def main():
    parser = argparse.ArgumentParser(description='Export analysis results')
    parser.add_argument('file', help='Path to the data file')
    parser.add_argument('--action', required=True,
                        choices=['excel', 'csv', 'html'],
                        help='Export format')
    parser.add_argument('--output', help='Output file path', required=True)
    parser.add_argument('--analysis-data', help='JSON analysis results data')
    parser.add_argument('--insights', help='JSON insights list')
    parser.add_argument('--charts', help='JSON charts list')
    parser.add_argument('--title', help='Report title', default='数据分析报告')
    parser.add_argument('--sheet', help='Excel sheet name', default=None)
    parser.add_argument('--sheet-name', help='Output Excel sheet name', default='分析结果')
    
    args = parser.parse_args()
    
    try:
        import pandas as pd
        from data_loader import load_file
        
        kwargs = {}
        if args.sheet:
            kwargs['sheet_name'] = args.sheet
        
        df, _ = load_file(args.file, **kwargs)
        
        # Parse JSON args
        charts = json.loads(args.charts) if args.charts else None
        analysis_results = json.loads(args.analysis_data) if args.analysis_data else None
        insights = json.loads(args.insights) if args.insights else None
        
        if args.action == 'excel':
            result = export_to_excel(df, args.output, args.sheet_name,
                                    charts=charts,
                                    analysis_results=analysis_results,
                                    insights=insights)
        elif args.action == 'csv':
            result = export_to_csv(df, args.output)
        elif args.action == 'html':
            result = generate_html_report(df, analysis_results, insights, charts,
                                         args.title, args.output)
        
        print(json.dumps(result, ensure_ascii=False, indent=2))
        
    except Exception as e:
        error_result = {'error': str(e), 'type': type(e).__name__}
        print(json.dumps(error_result, ensure_ascii=False, indent=2))
        sys.exit(1)


if __name__ == '__main__':
    main()
